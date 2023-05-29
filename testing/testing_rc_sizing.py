import openpyxl
  
# Call a Workbook() function of openpyxl 
# to create a new blank Workbook object
wb = openpyxl.Workbook()
  
# Get workbook active sheet  
# from the active attribute. 
sheet = wb.active
  
# writing to the specified cell
sheet.cell(row = 1, column = 1).value = ' hello '
  
sheet.cell(row = 2, column = 2).value = ' every'

sheet.row_dimensions[1].height = 48
  
# set the width of the column
#sheet.column_dimensions['B'].width = 5
  
# save the file
wb.save('dimension.xlsx')
