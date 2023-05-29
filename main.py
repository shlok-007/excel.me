import cv2
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from time import time

max_img_dim=222
rgb2hex = "#%02x%02x%02x"

imgName="meme1.jpeg"
img = cv2.imread("images\\"+imgName , cv2.IMREAD_UNCHANGED )

size=img.shape
mul_factor=max_img_dim/max(size[0:2])
new_dim=(int(size[1]*mul_factor),int(size[0]*mul_factor))

# if(mul_factor>1):
#     img=cv2.resize(img,new_dim,interpolation = cv2.INTER_CUBIC)
if(mul_factor<1):
    img=cv2.resize(img,new_dim,interpolation = cv2.INTER_AREA)

size=img.shape

print("Image Dimensions:",size[0:2])

def exelify(name):
    wb = Workbook()
    sheet = wb.active

    for i in range( 1 , size[0] + 1):
        sheet.row_dimensions[i].height = 48

    for rows in sheet.iter_rows(min_row=1, max_row=size[0], min_col=1, max_col=size[1]):
        for cell in rows:
            bgr = img[cell.row -1][cell.column -1]
            colour = "00" + (rgb2hex%(bgr[2] , bgr[1] , bgr[0]))[1:]
            cell.fill = PatternFill(start_color=colour, end_color=colour,fill_type = "solid")
    wb.save("XLs\\"+name)
if __name__ == "__main__":
    exelify("meme1_xl_"+str(max_img_dim)+".xlsx")
